from openpyxl.styles import Alignment
from openpyxl import Workbook, load_workbook
import io

#co jezeli nazwa firmy juz figuruje w liscie firm?
#lista_czynnikow, lista_wag, lista_ocen,
def licz_wyniki(lista_danych):
    for i in range(len(lista_danych[1])):
        if lista_danych[2][i] == None or lista_danych[3][i] == None:
            lista_danych[4][i] = None
        else:
            lista_danych[4][i] = round(lista_danych[2][i] * lista_danych[3][i],2)
   # return lista_danych

def full_import(filepath):
    with open(filepath, "rb") as f:
        file = io.BytesIO(f.read())
    wb = load_workbook(file, read_only=True)
    names = wb.sheetnames
    dane_firm = []
    for i in range(len(names)):
        firma = importer(wb,i)
        licz_wyniki(firma)
        dane_firm.append(firma)
    wb.close()
    del wb
    return dane_firm

def write_elements_to_sheet(lista_czynnikow, lista_wag, lista_ocen, lista_wynikow, nazwa_firmy, ws):
    ws.merge_cells('A1:D1')
    ws['A1'] = nazwa_firmy
    ws['A1'].alignment = Alignment(horizontal = "center")
    ws['A2'] = "Czynnik"
    ws['B2'] = "Waga"
    ws['C2'] = "Ocena"
    ws['D2'] = "Wynik"
    for row in range(3,len(lista_czynnikow)+3):
        ws.cell(column = 1, row = row, value = lista_czynnikow[row-3])
    for row in range(3,len(lista_wag)+3):
        ws.cell(column = 2, row = row, value = lista_wag[row-3])
    for row in range(3,len(lista_ocen)+3):
        ws.cell(column = 3, row = row, value = lista_ocen[row-3])
    for row in range(3,len(lista_wynikow)+3):
        ws.cell(column = 4, row = row, value = lista_wynikow[row-3])       
    return ws

def exporter(lista_czynnikow, lista_wag, lista_ocen, lista_wynikow, nazwa_firmy, liczba_firm):
    '''
    funkcja tworzy zeszyt z podanymi danymi firmy
    ostatni argument informuje o liczbie worksheetow z duplikowanymi danymi podanymi 
    w pozostałych argumentach
    input:
        lista_czynnikow - lista stringow
        lista_wag       - lista floatow
        lista_ocen      - lista integerow
        nazwa_firmy     - string
    return:
        Workbook: wb    - zeszyt Excel
    '''
    filename = "Dane.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = nazwa_firmy
    ws = write_elements_to_sheet(lista_czynnikow, lista_wag, lista_ocen, lista_wynikow,nazwa_firmy, ws)
    for i in range(1,liczba_firm):
            ws = wb.create_sheet(nazwa_firmy + str(i))
            ws = write_elements_to_sheet(lista_czynnikow, lista_wag, lista_ocen, lista_wynikow,nazwa_firmy, ws)
    wb.save(filename)
    wb.close()
    del wb
    print("zapisano")

def czytaj_skladniki_czynnikow(working_sheet, koordynaty_tytulowej_komorki):
    #zwraca czynninki i liczbe czynnikow
    lista_skladnikow = []
    column = koordynaty_tytulowej_komorki[1]
    row = koordynaty_tytulowej_komorki[0] + 1
    komorka = working_sheet.cell(row = row, column = column).value
    while komorka != None:
        lista_skladnikow.append(komorka)
        row = row + 1
        komorka = working_sheet.cell(row = row, column = column).value
    return lista_skladnikow, len(lista_skladnikow)

def czytaj_skladniki(working_sheet, koordynaty_tytulowej_komorki, liczba_czynnikow):
    lista_skladnikow = []
    column = koordynaty_tytulowej_komorki[1]
    row = koordynaty_tytulowej_komorki[0]+1
    for roww in range(row, row + liczba_czynnikow):
        lista_skladnikow.append(working_sheet.cell(row = roww, column = column).value)
    return lista_skladnikow

def importer(wb, nr_firmy): #nr_firmy od zera
    '''
    funkcja sczytuje dane na temat firmy z arkusza excel. przykladowy format dokumentu
    przedstawiam w pliku README.md
    input:
        nazwa_pliku - sciezka do arkusza ktory chcemy otworzyc
    return
        piecioelementowa lista list - wszystkie dane analizy z arkusza:

        nazwa_firmy, lista_czynnikow, lista_wag, lista_ocen, lista_wynikow
    '''
    #wb = load_workbook(filename = nazwa_pliku,read_only=True)
    #print("importuje")
    names = wb.sheetnames
    if nr_firmy >= len(names):
        raise ValueError("Niepoprawny nr firmy.")

    ws = wb[names[nr_firmy]]
    nazwa_firmy = ws['A1'].value
    czynnXY = []        # zmienne przechowujace wspolrzedne 
    wagiXY  = []        # tytulowych komorek kazdej kolumny
    ocenaXY = []       
    wynikiXY = []
    # szukanie tytulowych komorek
    for row in range (1, 6):
        for column in range (1, 6):
            komorka = ws.cell(row = row, column = column).value
            if komorka != None:
                komorka = str(komorka)
                komorka = komorka.strip().lower()
                if komorka.startswith("czynnik") and len(czynnXY) == 0:
                    czynnXY = [row, column]
                if komorka.startswith("wag") and len(wagiXY) == 0:
                    wagiXY = [row, column]
                if komorka.startswith("ocen") and len(ocenaXY) == 0:
                    ocenaXY = [row, column]
                if komorka.startswith("wyni") and len(wynikiXY) == 0:
                    wynikiXY = [row,column]
    # teoretycznie znalezione chyba ze ktos zrypal sprawe
    if len(wynikiXY) == 0:
        raise ValueError('nie przypisano wartosci do wynikow')

    lista_czynnikow, liczba_czynnikow = czytaj_skladniki_czynnikow(ws, czynnXY)
    lista_wag       = czytaj_skladniki(ws, wagiXY, liczba_czynnikow)
    lista_ocen      = czytaj_skladniki(ws, ocenaXY, liczba_czynnikow)
    lista_wynikow   = czytaj_skladniki(ws, wynikiXY, liczba_czynnikow)
    return nazwa_firmy, lista_czynnikow, lista_wag, lista_ocen, lista_wynikow

def daj_przykladowe_dane():
    '''
    ostatni zwracany argument to liczba firm w dokumencie przykładowym
    '''

    lista_ocen = [2,3,4]
    lista_wag = [0.2,0.3,0.5]
    lista_wynikow = [0.4,0.9,2]
    nazwa_firmy = "nazwa_firmy"
    lista_czynnikow = ["czynnik 1", "czynnik 2", "czynnik 3"]
    return lista_czynnikow, lista_wag, lista_ocen, lista_wynikow, nazwa_firmy, 2

def wykres_wynikow_firm(lista_firm): # lista tupli list. # potrzebna jeszcze nazwa firmy w liscie
    
    pass

